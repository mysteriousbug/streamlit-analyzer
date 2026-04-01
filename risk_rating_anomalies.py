import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import re

# ── CONFIG ──
INPUT_FILE = "eos.xlsx"
OUTPUT_FILE = "eos_comparison.xlsx"

# ── READ SHEETS ──
df_int = pd.read_excel(INPUT_FILE, sheet_name="internal")
df_ext = pd.read_excel(INPUT_FILE, sheet_name="external")

# ── NORMALIZE COLUMN NAMES ──
df_int.columns = df_int.columns.str.strip()
df_ext.columns = df_ext.columns.str.strip()

# ── IDENTIFY KEY COLUMNS ──
# Internal: "Product Name", "Product EOS Date Status"
# External: "Vendor", "Product", "Version/Release", "End of Life / End of Support"

int_name_col = [c for c in df_int.columns if "product" in c.lower() and "name" in c.lower()]
int_eos_col = [c for c in df_int.columns if "eos" in c.lower() or "end" in c.lower()]
int_vendor_col = [c for c in df_int.columns if "publisher" in c.lower()]
int_version_col = [c for c in df_int.columns if c.lower() == "software version"]

ext_product_col = [c for c in df_ext.columns if "product" in c.lower()]
ext_vendor_col = [c for c in df_ext.columns if "vendor" in c.lower()]
ext_version_col = [c for c in df_ext.columns if "version" in c.lower() or "release" in c.lower()]
ext_eos_col = [c for c in df_ext.columns if "end" in c.lower() or "eos" in c.lower()]

INT_NAME = int_name_col[0] if int_name_col else df_int.columns[0]
INT_EOS = int_eos_col[0] if int_eos_col else df_int.columns[1]
INT_VENDOR = int_vendor_col[0] if int_vendor_col else None
INT_VERSION = int_version_col[0] if int_version_col else None

EXT_PRODUCT = ext_product_col[0] if ext_product_col else df_ext.columns[1]
EXT_VENDOR = ext_vendor_col[0] if ext_vendor_col else df_ext.columns[0]
EXT_VERSION = ext_version_col[0] if ext_version_col else None
EXT_EOS = ext_eos_col[0] if ext_eos_col else df_ext.columns[-1]


def normalize(text):
    """Lowercase, strip, remove extra spaces and special chars for matching."""
    if pd.isna(text):
        return ""
    text = str(text).lower().strip()
    text = re.sub(r"[^\w\s.]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_date(val):
    """Parse various date formats into datetime."""
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val
    val = str(val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y", "%m-%d-%Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return None


# ── BUILD INTERNAL LOOKUP ──
# Key = (normalized_vendor + product, version) for precise matching
# Fallback = product name substring matching
internal_records = []
for _, row in df_int.iterrows():
    name_raw = str(row[INT_NAME]) if pd.notna(row[INT_NAME]) else ""
    # Parse "Tanium | Client | 7.4 | 7.4.9.1062-1 |" format
    parts = [p.strip().strip("|").strip() for p in name_raw.split("|") if p.strip()]
    vendor = parts[0] if len(parts) > 0 else ""
    product = parts[1] if len(parts) > 1 else parts[0] if parts else ""
    version = parts[2] if len(parts) > 2 else ""

    # Also pull from dedicated columns if available
    if INT_VENDOR and pd.notna(row.get(INT_VENDOR)):
        vendor = str(row[INT_VENDOR]).strip()
    if INT_VERSION and pd.notna(row.get(INT_VERSION)):
        version = str(row[INT_VERSION]).strip()

    internal_records.append({
        "raw_name": name_raw,
        "vendor": vendor,
        "product": product,
        "version": version,
        "vendor_norm": normalize(vendor),
        "product_norm": normalize(product),
        "version_norm": normalize(version),
        "eos_date": parse_date(row[INT_EOS]),
        "eos_raw": row[INT_EOS],
    })

# ── BUILD EXTERNAL LOOKUP ──
external_records = []
for _, row in df_ext.iterrows():
    vendor = str(row[EXT_VENDOR]).strip() if pd.notna(row[EXT_VENDOR]) else ""
    product = str(row[EXT_PRODUCT]).strip() if pd.notna(row[EXT_PRODUCT]) else ""
    version = str(row[EXT_VERSION]).strip() if EXT_VERSION and pd.notna(row.get(EXT_VERSION)) else ""

    external_records.append({
        "vendor": vendor,
        "product": product,
        "version": version,
        "vendor_norm": normalize(vendor),
        "product_norm": normalize(product),
        "version_norm": normalize(version),
        "eos_date": parse_date(row[EXT_EOS]),
        "eos_raw": row[EXT_EOS],
    })

# ── MATCHING LOGIC ──
results = []

for irec in internal_records:
    best_match = None
    best_score = 0  # 3 = vendor+product+version, 2 = vendor+product, 1 = product only

    for erec in external_records:
        score = 0
        # Check vendor match
        vendor_match = (
            irec["vendor_norm"] and erec["vendor_norm"]
            and (irec["vendor_norm"] in erec["vendor_norm"] or erec["vendor_norm"] in irec["vendor_norm"])
        )
        # Check product match
        product_match = (
            irec["product_norm"] and erec["product_norm"]
            and (irec["product_norm"] in erec["product_norm"] or erec["product_norm"] in irec["product_norm"])
        )
        # Check version match
        version_match = (
            irec["version_norm"] and erec["version_norm"]
            and (irec["version_norm"].startswith(erec["version_norm"])
                 or erec["version_norm"].startswith(irec["version_norm"])
                 or irec["version_norm"] == erec["version_norm"])
        )

        if vendor_match and product_match and version_match:
            score = 3
        elif vendor_match and product_match:
            score = 2
        elif product_match and vendor_match:
            score = 2
        elif product_match:
            score = 1

        if score > best_score:
            best_score = score
            best_match = erec

    if best_match and best_score >= 1:
        int_dt = irec["eos_date"]
        ext_dt = best_match["eos_date"]

        if int_dt and ext_dt:
            delta = (int_dt - ext_dt).days
            if delta == 0:
                status = "Match"
            elif delta > 0:
                status = f"Internal later by {delta} days"
            else:
                status = f"External later by {abs(delta)} days"
        else:
            status = "Date parse error"
            delta = None

        match_level = {3: "Vendor + Product + Version", 2: "Vendor + Product", 1: "Product Only"}[best_score]

        results.append({
            "Internal Product Name": irec["raw_name"],
            "Internal Vendor": irec["vendor"],
            "Internal Product": irec["product"],
            "Internal Version": irec["version"],
            "Internal EOS Date": int_dt.strftime("%Y-%m-%d") if int_dt else str(irec["eos_raw"]),
            "External Vendor": best_match["vendor"],
            "External Product": best_match["product"],
            "External Version": best_match["version"],
            "External EOS Date": ext_dt.strftime("%Y-%m-%d") if ext_dt else str(best_match["eos_raw"]),
            "Match Level": match_level,
            "Date Difference (days)": delta if delta is not None else "",
            "Status": status,
        })
    else:
        results.append({
            "Internal Product Name": irec["raw_name"],
            "Internal Vendor": irec["vendor"],
            "Internal Product": irec["product"],
            "Internal Version": irec["version"],
            "Internal EOS Date": irec["eos_date"].strftime("%Y-%m-%d") if irec["eos_date"] else str(irec["eos_raw"]),
            "External Vendor": "",
            "External Product": "",
            "External Version": "",
            "External EOS Date": "",
            "Match Level": "No Match Found",
            "Date Difference (days)": "",
            "Status": "No external match",
        })

df_result = pd.DataFrame(results)

# ── WRITE TO EXCEL WITH FORMATTING ──
wb = load_workbook(INPUT_FILE)

if "Comparison" in wb.sheetnames:
    del wb["Comparison"]
ws = wb.create_sheet("Comparison", 0)

# Styles
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
header_fill = PatternFill("solid", fgColor="2F5496")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

match_fill = PatternFill("solid", fgColor="C6EFCE")       # green
mismatch_fill = PatternFill("solid", fgColor="FFC7CE")     # red
no_match_fill = PatternFill("solid", fgColor="FFEB9C")     # yellow
alt_row_fill = PatternFill("solid", fgColor="D6E4F0")      # light blue

match_font = Font(color="006100", name="Arial", size=10)
mismatch_font = Font(color="9C0006", name="Arial", size=10)
no_match_font = Font(color="9C6500", name="Arial", size=10)
default_font = Font(name="Arial", size=10)

# Write headers
headers = list(df_result.columns)
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Write data
for row_idx, row_data in enumerate(df_result.itertuples(index=False), 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = default_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        if row_idx % 2 == 0:
            cell.fill = alt_row_fill

    # Color-code the Status column
    status_col = headers.index("Status") + 1
    status_cell = ws.cell(row=row_idx, column=status_col)
    status_val = str(status_cell.value)

    if status_val == "Match":
        status_cell.fill = match_fill
        status_cell.font = match_font
    elif "No" in status_val:
        status_cell.fill = no_match_fill
        status_cell.font = no_match_font
    elif "later" in status_val:
        status_cell.fill = mismatch_fill
        status_cell.font = mismatch_font

# Auto-width columns
for col_idx in range(1, len(headers) + 1):
    max_len = len(str(headers[col_idx - 1]))
    for row_idx in range(2, len(df_result) + 2):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val:
            max_len = max(max_len, len(str(val)))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

# Freeze top row
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_FILE)

# ── SUMMARY ──
total = len(results)
matched = sum(1 for r in results if r["Status"] != "No external match")
date_match = sum(1 for r in results if r["Status"] == "Match")
date_mismatch = sum(1 for r in results if "later" in str(r["Status"]))
no_ext = total - matched

print(f"\n{'='*50}")
print(f"EOS COMPARISON COMPLETE")
print(f"{'='*50}")
print(f"Total internal products:  {total}")
print(f"Matched to external:      {matched}")
print(f"  - EOS dates match:      {date_match}")
print(f"  - EOS dates differ:     {date_mismatch}")
print(f"No external match found:  {no_ext}")
print(f"\nOutput: {OUTPUT_FILE}")
